from django.shortcuts import render
from django.shortcuts import render
from openpyxl import load_workbook
from .models import Product


def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            #name, price, quantity = row
            name=row
            #Product.objects.create(name=name, price=price, quantity=quantity)
            Product.objects.create(name=name)

        return render(request, 'import.html')

    return render(request, 'import.html')
 